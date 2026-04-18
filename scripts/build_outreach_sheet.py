"""One-shot generator for the self-heal customer-discovery tracking sheet.

Writes `C:/Users/Johin/Desktop/self_heal_outreach.xlsx`. Idempotent.

Data is populated from public-source research (engineering blogs, team
pages, job ads, published talks). Every contact name is a STARTING
POINT that must be verified on LinkedIn before any DM is sent. People
change roles; titles drift; companies get acquired.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path("C:/Users/Johin/Desktop/self_heal_outreach.xlsx")


# (rank, company, fit_notes, starter_contacts, linkedin_search_hint, risks,
#  first_name_placeholder, personalization_hook, x_handles)
# starter_contacts: list of (name, title_at_time_of_research, source)
# x_handles: list of (label, @handle) — first entry is the company handle
#            when found; individual handles follow.
TARGETS: list[tuple] = [
    (
        1,
        "Harvey",
        "Legal AI agents in production, enterprise buyers, $11B valuation Mar 2026",
        [
            ("Joey Wang", "Engineering Lead, background agents", "harvey.ai/blog/author/joey-wang"),
        ],
        'site:linkedin.com "Harvey AI" ("platform" OR "infrastructure" OR "agent")',
        "Large eng team; find the sub-team building agents specifically",
        "Joey",
        "saw your work on background agents at Harvey",
        [
            ("Harvey (company)", "@harvey__ai"),
            ("Joey Wang", "@ZongZiWang"),
        ],
    ),
    (
        2,
        "Sierra",
        "Customer-service agents, Bret Taylor/Clay Bavor team, enterprise sales motion",
        [
            ("Arya Asemanfar", "Head of Product Engineering", "freeplay.ai blog"),
        ],
        'site:linkedin.com "Sierra" ("agent platform" OR "agent infrastructure")',
        "Bret Taylor halo; senior folks get heavy inbound",
        "Arya",
        "your Freeplay interview on enterprise agent reliability landed exactly where we're building",
        [
            ("Sierra (company)", "@SierraPlatform"),
            ("Arya Asemanfar", "@a_a"),
        ],
    ),
    (
        3,
        "Decagon",
        "Customer-service agents, large eng manager role published",
        [
            ("Cyrus Asgari", "Lead Research Engineer, fine-tuning", "decagon.ai"),
        ],
        'site:linkedin.com Decagon ("agent" OR "platform" OR "infra")',
        "Research-leaning team; frame as reliability/eval not just repair",
        "Cyrus",
        "the rapid iteration loop across the agent lifecycle Decagon ships is where runtime repair would slot in",
        [
            ("Decagon (company)", "@DecagonAI"),
            ("Cyrus Asgari", "@cyrus_asgari"),
        ],
    ),
    (
        4,
        "Cresta",
        "Contact-center AI agents, talent-dense eng team",
        [
            ("(none found publicly)", "try: Director of Engineering or VP AI", "cresta.com/careers"),
        ],
        'site:linkedin.com Cresta ("director engineering" OR "VP engineering")',
        "Acquired Grey Parrot; mid-size",
        "{FirstName}",
        "real-time contact-center agents have one of the sharpest \"when they fail, someone hears it\" loops in production today",
        [
            ("Cresta (company)", "@cresta"),
        ],
    ),
    (
        5,
        "Glean",
        "Enterprise search + agents, named eng blog authors",
        [
            ("Neil Dhruva", "Engineering (next-gen prompting blog author)", "glean.com/blog"),
            ("Eddie Zhou", "Engineering (context engineering blog author)", "glean.com/blog"),
            ("Arjun Landes", "Engineering (context engineering blog author)", "glean.com/blog"),
        ],
        'site:linkedin.com Glean ("platform" OR "AI platform" OR "context")',
        "Big co now; long procurement",
        "Eddie",
        "your \"context engineering shifts to the platform\" post was exactly the right framing",
        [
            ("Glean (company)", "@gleanwork"),
            ("Eddie Zhou", "@eddiedzhou"),
            ("Arjun Landes", "@arjunlandes"),
        ],
    ),
    (
        6,
        "Writer",
        "Enterprise gen AI, agentic playbooks launched Mar 2026, compliance-conscious buyers",
        [
            ("(none found publicly)", "target: VP Eng or Head of AI Platform", "writer.com/company/about"),
        ],
        'site:linkedin.com writer.com ("platform engineer" OR "AI platform")',
        "Very enterprise-sales-focused; eng team may be less accessible",
        "{FirstName}",
        "Agent Skills + Playbooks for business teams is exactly where audit-and-policy stories start to matter",
        [
            ("Writer (company)", "@Get_Writer"),
        ],
    ),
    (
        7,
        "Cognition (Devin + Windsurf)",
        "Coding agent, acquired Codeium/Windsurf Dec 2025 for ~$250M — now the combined target (previously #11)",
        [
            ("Scott Wu", "Co-founder/CEO", "public"),
            ("Walden Yan", "Co-founder", "public"),
            ("Varun Mohan", "CEO of Windsurf (inside Cognition post-acq)", "public"),
        ],
        'site:linkedin.com Cognition (Devin OR Windsurf OR "agent infra")',
        "Cognition + Windsurf is one company now; pick one senior contact, don't double-up",
        "Scott",
        "the Devin 2025 performance review was one of the most honest looks at agent failure modes I've read — and the Windsurf acquisition consolidates exactly the category we're building for",
        [
            ("Scott Wu", "@ScottWu46"),
            ("Walden Yan", "@walden_yan"),
            ("Varun Mohan (Windsurf, now inside Cognition)", "@_mohansolo"),
            ("Codeium/Windsurf (product)", "@codeiumdev"),
        ],
    ),
    (
        8,
        "Factory",
        "Agent-native dev platform, $150M Series C, enterprise customers (MongoDB, E&Y)",
        [
            ("(team ex-Nuro/Glean/Scale/MongoDB)", "target: Head of Platform or Head of Agent Infra", "factory.ai/company"),
        ],
        'site:linkedin.com "Factory AI" ("platform" OR "droid" OR "agent")',
        "Competitive adjacency: they also ship a platform; pitch as complementary",
        "{FirstName}",
        "self-healing builds is literally in Factory's pitch; curious how you think about the runtime-repair layer vs the CI layer",
        [
            ("Factory (company)", "@FactoryAI"),
        ],
    ),
    (
        9,
        "Magic.dev",
        "Coding agents, $515M raised, custom ML stack, small secretive team",
        [
            ("(secretive)", "co-founders Eric Steinberger, Sebastian De Ro", "magic.dev"),
        ],
        'site:linkedin.com magic.dev',
        "SKIP unless warm intro; no public platform-eng folks",
        "{FirstName}",
        "the 100M-token LTM Net approach is singular; framing repair/eval at that context scale is interesting",
        [
            ("(no company X handle found)", ""),
        ],
    ),
    (
        10,
        "Cursor / Anysphere",
        "Code editor agents, $29B valuation, 35% of internal PRs from autonomous agents",
        [
            ("Michael Truell", "Co-founder/CEO", "founder / public"),
            ("Sualeh Asif", "Co-founder", "founder / public"),
            ("Aman Sanger", "Co-founder", "founder / public"),
        ],
        'site:linkedin.com "Anysphere" ("platform" OR "infra")',
        "Founders are high-signal but swamped with inbound; target platform ICs instead",
        "{FirstName}",
        "35% of your PRs from autonomous agents is exactly the scale where runtime reliability tooling starts to pay for itself",
        [
            ("Cursor (company)", "@cursor_ai"),
            ("Michael Truell", "@mntruell"),
            ("Sualeh Asif", "@sualehasif996"),
            ("Aman Sanger", "@amanrsanger"),
        ],
    ),
    (
        11,
        "(REMOVED: now merged with #7 Cognition)",
        "Codeium/Windsurf acquired by Cognition Dec 2025; consolidated into target #7. Use slot #11 for a replacement target if you want to keep 20.",
        [
            ("n/a", "merged into #7", ""),
        ],
        "n/a",
        "Do not DM twice at the same combined company.",
        "",
        "",
        [
            ("(see target #7)", ""),
        ],
    ),
    (
        12,
        "Lindy",
        "General-purpose agent platform",
        [
            ("Luiz Scheidegger", "Head of Engineering", "public role"),
        ],
        'site:linkedin.com Lindy ("engineering" OR "agent")',
        "SMB-focused customer base; Team tier likely better fit than Enterprise",
        "Luiz",
        "the \"AI employee\" framing at Lindy lives or dies on production reliability",
        [
            ("Lindy (company)", "@Lindychii"),
            ("Luiz Scheidegger", "@luizscheidegger"),
        ],
    ),
    (
        13,
        "Browserbase",
        "Browser agent infrastructure, customers include Perplexity, Vercel",
        [
            ("Paul Klein IV", "Founder/CEO", "public"),
            ("Dominic Saadi", "Founding Engineer", "Contrary Research"),
            ("Walker Griggs", "Technical Lead", "Contrary Research"),
        ],
        'site:linkedin.com Browserbase ("engineer" OR "infra")',
        "Infra-focused company; natural partnership potential",
        "Paul",
        "browser agents are the most failure-prone runtime in production today; natural home for a repair layer",
        [
            ("Browserbase (company)", "@browserbasehq"),
            ("Paul Klein IV", "@pk_iv"),
            ("Walker Griggs", "@WebDevScaresMe"),
        ],
    ),
    (
        14,
        "Arcade",
        "Agent tools platform, auth/tooling focus, $12M seed",
        [
            ("Jo Stevens", "AI/MLOps Engineer", "arcade.dev"),
            ("Pascal Matthiesen", "Distributed Systems Engineer", "arcade.dev"),
        ],
        'site:linkedin.com "Arcade.dev" ("engineer" OR "platform")',
        "Smaller co; Team tier fit; possible partner (LangChain integration)",
        "Jo",
        "your tool-calling runtime sits one layer below where repair fires; curious whether they compose",
        [
            ("(no company X handle confirmed; try @ArcadeAI)", ""),
        ],
    ),
    (
        15,
        "Replit",
        "Replit Agent in production, active Agent Platform hiring",
        [
            ("(actively hiring)", "EM Agent Infrastructure + Staff SWE Agent Platform open", "jobs.ashbyhq.com/replit"),
        ],
        'site:linkedin.com Replit ("agent platform" OR "agent infrastructure")',
        "Strong current demand (hiring); cold DM to current agent-platform ICs",
        "{FirstName}",
        "your open EM Agent Infrastructure req is a tell that reliability is a live problem right now",
        [
            ("Replit (company)", "@Replit"),
        ],
    ),
    (
        16,
        "Perplexity",
        "Agentic search, Engineering Manager Agents role open, $250-325k comp",
        [
            ("(actively hiring)", "EM Agents + AI Infra Engineer open", "jobs.ashbyhq.com/perplexity"),
        ],
        'site:linkedin.com Perplexity ("agents" OR "agent platform")',
        "High inbound volume; require strong personalization",
        "{FirstName}",
        "the EM Agents posting at Perplexity is a tell that agentic reliability is top-of-mind",
        [
            ("Perplexity (company)", "@perplexity_ai"),
        ],
    ),
    (
        17,
        "Vercel (v0 team)",
        "v0 agent workflows launched for 2026, enterprise AI Cloud positioning",
        [
            ("(no specific platform-lead found)", "target: Head of v0 Engineering or VP Engineering AI", "vercel.com/blog/introducing-the-new-v0"),
        ],
        'site:linkedin.com Vercel ("v0" OR "AI platform")',
        "Dev-tools buyer; understand OSS; good fit for partnership not just sale",
        "{FirstName}",
        "v0's new end-to-end agent workflows on Vercel's self-driving infra is exactly where a runtime repair layer fits",
        [
            ("Vercel (company)", "@vercel"),
        ],
    ),
    (
        18,
        "Retool",
        "Retool Agents platform, enterprise-grade infra, JS/TS/React stack",
        [
            ("(small AI team inside Retool)", "target: Eng Manager Agents or Head of AI Product", "retool.com/agents"),
        ],
        'site:linkedin.com Retool ("agents" OR "AI agent")',
        "Internal-tools-first buyer; agent use case still maturing there",
        "{FirstName}",
        "Retool Agents with built-in governance + evaluations + observability is the reference shape of the platform story",
        [
            ("Retool (company)", "@retool"),
        ],
    ),
    (
        19,
        "Moveworks",
        "Employee-service agents, acquired by ServiceNow Dec 2025",
        [
            ("Chang Liu", "Director of Engineering, ML", "theorg.com"),
            ("Mayank Khanwalker", "Director of Engineering (Bengaluru)", "Moveworks press"),
        ],
        'site:linkedin.com Moveworks ("platform" OR "core platform" OR "agentic")',
        "Now part of ServiceNow; budget owners changed; may need longer sales cycle",
        "Chang",
        "curious how the Reasoning Engine plus ServiceNow integration surfaces reliability controls at enterprise scale",
        [
            ("(no handles confirmed; try @moveworks, @servicenow)", ""),
        ],
    ),
    (
        20,
        "Adept (archive)",
        "Agent foundation models — team largely absorbed by Amazon 2024",
        [
            ("(team dispersed)", "low priority, may find ex-Adept engs now at other targets", "public"),
        ],
        "n/a",
        "SKIP in current form; use the LinkedIn search only to find ex-Adept engs now at other target companies",
        "",
        "",
        [
            ("(skip)", ""),
        ],
    ),
]


def _compose_dm(first_name: str, company: str, hook: str) -> str:
    """Build a ready-to-send LinkedIn DM for a target.

    Under 2000 chars (InMail limit). If the first_name contains a
    placeholder like {FirstName} the user must fill it in before sending.
    Returns "" for targets we're skipping (rank 20 Adept).
    """
    if not hook:
        return ""
    return (
        f"Hi {first_name}, {hook}.\n\n"
        f"I'm building self-heal (github.com/Johin2/self-heal), an OSS library "
        f"that makes LLM agents repair themselves at runtime. Doing customer "
        f"discovery at {company} and companies like yours before building the "
        f"hosted control plane (audit log, policy engine, observability).\n\n"
        f"20 min to hear how your team handles agent failures in production today? "
        f"Happy with a \"not a fit\" reply too. Mid-week this week or next?\n\n"
        f"Johin"
    )


TEMPLATES = {
    "LinkedIn connection request note (300 chars max, free-tier safe)": (
        "Hi {FirstName}, building self-heal (github.com/Johin2/self-heal) — OSS "
        "that makes LLM agents repair themselves at runtime. Doing customer "
        "discovery before the hosted layer. 20 min to hear how your team "
        "handles agent failures? Happy with a \"not a fit\" reply."
    ),
    "LinkedIn DM (after connection accepts)": (
        "Hi {FirstName}, thanks for connecting. Quick context: I'm building "
        "self-heal (github.com/Johin2/self-heal), an OSS library that makes LLM "
        "agents repair themselves at runtime. Doing customer discovery at "
        "{Company} and similar places before building the hosted control plane "
        "(audit log, policy engine, observability).\n\n"
        "20 min to hear how your team handles agent failures in prod today? "
        'Happy with "not a fit" too. Mid-week this week or next?\n\n'
        "Johin"
    ),
    "Cold email (primary recommended channel if you don't have Premium)": (
        "Subject: 20 min on how {Company} handles agent failures in prod?\n\n"
        "Hi {FirstName},\n\n"
        "{one-line personalization hook from the Targets tab}.\n\n"
        "I'm building self-heal (github.com/Johin2/self-heal), an OSS library "
        "that makes LLM agents repair themselves at runtime. Before I build the "
        "hosted control plane (audit log, policy engine, observability), I'm "
        "doing customer discovery at ~20 places running agents in production — "
        "{Company} is one of them.\n\n"
        "20 min to hear how your team handles agent failures today? Mid-week "
        'this week or next works, and "not a fit" is also a useful reply.\n\n'
        "Johin Johny\n"
        "github.com/Johin2/self-heal"
    ),
    "X / Twitter DM (short, casual)": (
        "Hey {FirstName} — {one-line hook about their recent tweet/post}. "
        "Building self-heal (OSS agent-repair library, github.com/Johin2/self-heal) "
        "and doing customer discovery before the hosted layer. 20 min to hear "
        "how your team handles agent failures? Totally fine to pass."
    ),
    "Follow-up (day 7 if no reply, any channel)": (
        "Bumping this in case it got buried. If now is not a fit, no worries, "
        'a one-line "not now" is also useful signal. Thanks either way.'
    ),
    "Personalization add-on (optional one line above the pitch)": (
        "Enjoyed your recent {post / talk / feature} on {specific thing}."
    ),
    "Call agenda (20 minutes)": (
        "0-2 min: thanks, set expectation (I mostly want to listen)\n"
        "2-10 min: tell me about the agents you run in prod and where they fail\n"
        "10-15 min: when an agent fails in prod today, what happens? Who's paged? "
        "What's the remediation path? What would you pay to make that better?\n"
        "15-18 min: 90-second pitch + 15s GIF, then \"does this resonate?\"\n"
        "18-20 min: ask for referrals (who else at {Company} or in your network?)"
    ),
    "Free-tier workflow (no LinkedIn Premium)": (
        "1. Cold email (primary) — Apollo.io free tier gives 60 emails/mo. Use the "
        "Cold Email template above. 25-40% response rate from engineers.\n"
        "2. X/Twitter DM (warmup + secondary) — follow the target, like/reply to "
        "a recent tweet for a day or two, THEN DM. Use the X DM template.\n"
        "3. LinkedIn connection request (tertiary) — 300-char note, 3-5/day max to "
        "avoid spam flags. Once they accept, you can DM freely on free tier."
    ),
}


READ_ME = [
    ("Purpose", "Track self-heal customer-discovery outreach (Phase 0 of platform plan)."),
    ("Target volume", "20 personalized DMs. Checkpoint 1: >= 5 booked calls."),
    ("Pace", "3-5 DMs/day to avoid LinkedIn anti-spam flags. Spread over ~week 1."),
    ("Kill criterion", "< 5 calls booked after 20 DMs, or < 2 express willingness to pay."),
    ("Names source", "ALL contact names are from PUBLIC sources (blogs, team pages, talks, job ads). VERIFY on LinkedIn before DMing. People change roles; this is a starting point, not a vetted list."),
    ("Avoid", "Mass-DMing the same template to 20 people — LinkedIn flags this. Personalize each."),
    ("Stack", "Google Sheets works fine if you prefer — this xlsx is just a starting template."),
    ("Plan file", "C:/Users/Johin/.claude/plans/delightful-marinating-cat.md"),
]


# -------- xlsx construction --------


def _bold(cell) -> None:
    cell.font = Font(bold=True)


def _header_fill(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    _bold(cell)


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "self-heal customer discovery outreach"
    ws["A1"].font = Font(size=14, bold=True)
    for i, (k, v) in enumerate(READ_ME, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    for row_idx in range(3, 3 + len(READ_ME)):
        ws.row_dimensions[row_idx].height = 35


def build_targets(wb: Workbook) -> None:
    ws = wb.create_sheet("Targets")
    headers = [
        "Rank",
        "Company",
        "Fit notes",
        "Starter contact(s)  (VERIFY on LinkedIn)",
        "Suggested LinkedIn search",
        "Risks / notes",
        "X / Twitter handles  (VERIFY; accounts change)",
        "Draft DM  (edit {FirstName} where needed, verify contact first)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _header_fill(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, row in enumerate(TARGETS, start=2):
        (
            rank,
            company,
            fit,
            contacts,
            search,
            risks,
            first_name,
            hook,
            x_handles,
        ) = row
        contact_str = "\n".join(
            f"- {name}  |  {title}  |  src: {src}" for name, title, src in contacts
        )
        x_str = "\n".join(
            f"- {label}: {handle}" if handle else f"- {label}"
            for label, handle in x_handles
        )
        dm = _compose_dm(first_name, company, hook)
        ws.cell(row=row_idx, column=1, value=rank)
        ws.cell(row=row_idx, column=2, value=company).font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=fit)
        ws.cell(row=row_idx, column=4, value=contact_str)
        ws.cell(row=row_idx, column=5, value=search)
        ws.cell(row=row_idx, column=6, value=risks)
        ws.cell(row=row_idx, column=7, value=x_str)
        ws.cell(row=row_idx, column=8, value=dm or "(skip this target)")
        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        ws.row_dimensions[row_idx].height = max(
            160, 18 * (len(contacts) + len(x_handles) + 1)
        )

    widths = [6, 22, 42, 50, 42, 38, 45, 75]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


OUTREACH_LOG_ENTRIES: list[dict] = [
    # Sent on 2026-04-18 via X DM. Channel noted in `notes` field.
    {
        "date_sent": "2026-04-18",
        "name": "Joey Wang",
        "company": "Harvey",
        "role": "Engineering Lead, background agents",
        "linkedin_url": "https://linkedin.com/in/zongziwang/",
        "personalization_hook": "saw your work on background agents at Harvey",
        "reply_status": "no_reply",
        "notes": "channel: X DM @ZongZiWang",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Walker Griggs",
        "company": "Browserbase",
        "role": "Technical Lead, Core Team",
        "linkedin_url": "https://linkedin.com/in/walkergriggs/",
        "personalization_hook": "browser agents are the most failure-prone runtime today",
        "reply_status": "no_reply",
        "notes": "channel: X DM @WebDevScaresMe",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Paul Klein IV",
        "company": "Browserbase",
        "role": "Founder / CEO",
        "linkedin_url": "https://linkedin.com/in/paulkleiniv",
        "personalization_hook": "browser agents are the most failure-prone runtime today",
        "reply_status": "no_reply",
        "notes": "channel: X DM @pk_iv  |  double-tap on Browserbase with Walker; if either replies, don't bother the other",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Michael Truell",
        "company": "Cursor / Anysphere",
        "role": "Co-founder / CEO",
        "linkedin_url": "",
        "personalization_hook": "35% of your PRs from autonomous agents is exactly the scale where runtime reliability tooling starts to pay for itself",
        "reply_status": "no_reply",
        "notes": "channel: X DM @mntruell  |  CEO inbox is very noisy; low response rate expected",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Sualeh Asif",
        "company": "Cursor / Anysphere",
        "role": "Co-founder / CPO",
        "linkedin_url": "https://linkedin.com/in/sualeh-a-1a4a97116/",
        "personalization_hook": "35% of your PRs from autonomous agents is exactly the scale where runtime reliability tooling starts to pay for itself",
        "reply_status": "no_reply",
        "notes": "channel: X DM @sualehasif996  |  triple-tap on Cursor founders; if any one replies, drop the others",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Aman Sanger",
        "company": "Cursor / Anysphere",
        "role": "Co-founder / COO",
        "linkedin_url": "https://linkedin.com/in/aman-sanger-482243171/",
        "personalization_hook": "35% of your PRs from autonomous agents is exactly the scale where runtime reliability tooling starts to pay for itself",
        "reply_status": "no_reply",
        "notes": "channel: X DM @amanrsanger  |  triple-tap on Cursor founders",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Eddie Zhou",
        "company": "Glean",
        "role": "Founding Engineer",
        "linkedin_url": "https://linkedin.com/in/eddie-zhou-70364b54",
        "personalization_hook": "your \"context engineering shifts to the platform\" post was exactly the right framing",
        "reply_status": "no_reply",
        "notes": "channel: X DM @eddiedzhou",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Arjun Landes",
        "company": "Glean",
        "role": "Engineering Manager, Search Ranking",
        "linkedin_url": "https://linkedin.com/in/arjunlandes/",
        "personalization_hook": "your \"context engineering shifts to the platform\" post was exactly the right framing",
        "reply_status": "no_reply",
        "notes": "channel: X DM @arjunlandes  |  double-tap on Glean with Eddie",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Scott Wu",
        "company": "Cognition (Devin)",
        "role": "Co-founder / CEO",
        "linkedin_url": "",
        "personalization_hook": "the Devin 2025 performance review post was one of the most honest looks at agent failure modes I've read",
        "reply_status": "no_reply",
        "notes": "channel: X DM @ScottWu46",
    },
    {
        "date_sent": "2026-04-18",
        "name": "Walden Yan",
        "company": "Cognition (Devin)",
        "role": "Co-founder",
        "linkedin_url": "",
        "personalization_hook": "the Devin 2025 performance review post was one of the most honest looks at agent failure modes I've read",
        "reply_status": "no_reply",
        "notes": "channel: X DM @walden_yan  |  double-tap on Cognition with Scott",
    },
    {
        "date_sent": "",
        "name": "Varun Mohan",
        "company": "Cognition (Windsurf, post-acquisition)",
        "role": "CEO of Windsurf (now inside Cognition)",
        "linkedin_url": "https://linkedin.com/in/varunkmohan",
        "personalization_hook": "Latent Space episode on the enterprise IDE stack + Cognition acquisition consolidates exactly the category we're building for",
        "reply_status": "draft",
        "notes": "channel: X DM @_mohansolo  |  DRAFT NOT YET SENT. Finish and send, but after deciding whether to avoid double-tapping with Scott/Walden (same company now)",
    },
]


def build_outreach_log(wb: Workbook) -> None:
    ws = wb.create_sheet("Outreach Log")
    headers = [
        "date_sent",
        "name",
        "company",
        "role",
        "linkedin_url",
        "personalization_hook",
        "reply_status",
        "date_replied",
        "call_date",
        "willing_to_pay",
        "pain_level_1_5",
        "notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _header_fill(ws.cell(row=1, column=col_idx, value=h))

    # Legend row
    legend = [
        "YYYY-MM-DD",
        "full name",
        "company",
        "title",
        "profile URL",
        "one specific thing or blank",
        "no_reply | replied | declined | booked | completed | not_a_fit | draft",
        "YYYY-MM-DD or blank",
        "YYYY-MM-DD or blank",
        "yes | no | unknown | maybe",
        "1=no pain, 5=actively firefighting",
        "free text (include channel: X | LinkedIn | email | connection)",
    ]
    for col_idx, v in enumerate(legend, start=1):
        cell = ws.cell(row=2, column=col_idx, value=v)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = Alignment(wrap_text=True)

    # Prefilled rows from the user's in-flight outreach.
    for row_idx, entry in enumerate(OUTREACH_LOG_ENTRIES, start=3):
        ws.cell(row=row_idx, column=1, value=entry.get("date_sent", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("company", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("role", ""))
        ws.cell(row=row_idx, column=5, value=entry.get("linkedin_url", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("personalization_hook", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("reply_status", "no_reply"))
        ws.cell(row=row_idx, column=8, value=entry.get("date_replied", ""))
        ws.cell(row=row_idx, column=9, value=entry.get("call_date", ""))
        ws.cell(row=row_idx, column=10, value=entry.get("willing_to_pay", ""))
        ws.cell(row=row_idx, column=11, value=entry.get("pain_level_1_5", ""))
        ws.cell(row=row_idx, column=12, value=entry.get("notes", ""))
        for col_idx in range(1, 13):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        ws.row_dimensions[row_idx].height = 45

    widths = [11, 20, 22, 32, 42, 40, 14, 12, 12, 14, 14, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def build_templates(wb: Workbook) -> None:
    ws = wb.create_sheet("Templates")
    ws["A1"] = "Outreach templates"
    ws["A1"].font = Font(size=14, bold=True)

    row = 3
    for title, body in TEMPLATES.items():
        cell = ws.cell(row=row, column=1, value=title)
        _bold(cell)
        row += 1
        body_cell = ws.cell(row=row, column=1, value=body)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 20 * (body.count("\n") + 3)
        row += 2

    ws.column_dimensions["A"].width = 110


def main() -> None:
    wb = Workbook()
    build_readme(wb)
    build_targets(wb)
    build_outreach_log(wb)
    build_templates(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
